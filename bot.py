from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
import openpyxl
import os
import json
import re
import requests
import google.generativeai as genai
from datetime import datetime

app = Flask(__name__)
EXCEL_FILE = "orders.xlsx"
CONTACTS_FILE = "contacts.json"
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")

genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel("gemini-1.5-flash")

def setup_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(["Order No", "Date", "Time", "Sender Name", "Phone", "Item", "Qty (kg)", "Qty (pieces/nag)", "Qty (bundles/gaddi)"])
        wb.save(EXCEL_FILE)

def load_contacts():
    try:
        if os.path.exists(CONTACTS_FILE):
            with open(CONTACTS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except:
        pass
    return {}

def get_name(phone):
    contacts = load_contacts()
    phone_clean = phone.strip().replace(" ", "")
    for key in contacts:
        if key.strip().replace(" ", "") == phone_clean:
            return contacts[key]
    return phone

def translate_item(item):
    try:
        prompt = f"Translate this vegetable/fruit/food item name to English. Return ONLY the English name, nothing else: '{item}'"
        response = model.generate_content(prompt)
        return response.text.strip()
    except:
        return item

def parse_quantity(qty_str):
    qty_str = qty_str.strip().lower()
    gm_match = re.search(r'(\d+\.?\d*)\s*(gm|gram|grm|g)\b', qty_str)
    if gm_match:
        grams = float(gm_match.group(1))
        return str(round(grams / 1000, 3)), "", ""
    kg_match = re.search(r'(\d+\.?\d*)\s*(kg|kilo)\b', qty_str)
    if kg_match:
        return kg_match.group(1), "", ""
    nag_match = re.search(r'(\d+\.?\d*)\s*(nag|nug|piece|pcs|pc|nos|no)\b', qty_str)
    if nag_match:
        return "", nag_match.group(1), ""
    gaddi_match = re.search(r'(\d+\.?\d*)\s*(gaddi|gadi|bundle|bunch)\b', qty_str)
    if gaddi_match:
        return "", "", gaddi_match.group(1)
    num_match = re.search(r'(\d+\.?\d*)', qty_str)
    if num_match:
        return num_match.group(1), "", ""
    return "", "", ""

def log_order(sender_name, phone, item, kg, pieces, bundles):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    order_no = ws.max_row
    date = datetime.now().strftime("%d-%m-%Y")
    time_now = datetime.now().strftime("%H:%M:%S")
    ws.append([order_no, date, time_now, sender_name, phone, item, kg, pieces, bundles])
    wb.save(EXCEL_FILE)
    print(f"✅ Order! | {sender_name} | {item} | kg:{kg} | pcs:{pieces} | bundle:{bundles}")

def extract_orders_from_image(image_url, account_sid, auth_token):
    try:
        response = requests.get(image_url, auth=(account_sid, auth_token))
        image_data = response.content
        prompt = """This image contains a list of vegetable/fruit orders written in Hindi or English or mixed.
        Extract all items and their quantities. Convert all item names to English.
        Return ONLY in this exact format, one item per line:
        ITEM_NAME | QUANTITY | UNIT
        Units should be: kg, gm, nag, gaddi
        Example:
        Potato | 10 | kg
        Coriander | 500 | gm
        Lemon | 2 | nag
        Mint | 1 | gaddi"""
        image_part = {"mime_type": "image/jpeg", "data": image_data}
        response = model.generate_content([prompt, image_part])
        return response.text.strip()
    except Exception as e:
        print(f"Image error: {e}")
        return ""

def parse_gemini_image_response(text):
    orders = []
    lines = text.strip().split("\n")
    for line in lines:
        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3:
                item = parts[0].strip()
                quantity = parts[1].strip()
                unit = parts[2].strip().lower()
                if unit == "gm":
                    kg = str(round(float(quantity) / 1000, 3))
                    orders.append((item, kg, "", ""))
                elif unit == "kg":
                    orders.append((item, quantity, "", ""))
                elif unit == "nag":
                    orders.append((item, "", quantity, ""))
                elif unit == "gaddi":
                    orders.append((item, "", "", quantity))
                else:
                    orders.append((item, quantity, "", ""))
    return orders

def parse_line(line):
    line = line.strip()
    if not line:
        return None
    line = re.sub(r'^\d+[\)\.:\-]\s*', '', line)
    line_lower = line.lower()

    # Format 1: "item ke liye order: qty"
    if "order:" in line_lower:
        parts = line_lower.split("order:")
        item = parts[0].replace("ke liye", "").strip()
        quantity = parts[1].strip()
        if item and quantity:
            return (item, quantity)

    # Format 2: "item - qty" or "item: qty" or "item = qty"
    match = re.match(r'^([\w\s\u0900-\u097F]+?)\s*[\-:=]+\s*([\d\w\s\.]+)$', line)
    if match:
        return (match.group(1).strip(), match.group(2).strip())

    # Format 3: "40 kanda" — number pehle space ke saath
    match = re.match(r'^(\d+\.?\d*)\s+([\w\u0900-\u097F][\w\u0900-\u097F\s]*)$', line)
    if match:
        return (match.group(2).strip(), match.group(1).strip())

    # Format 4: "20tamatar" — number aur text sath mein
    match = re.match(r'^(\d+)([\u0900-\u097Fa-zA-Z][a-zA-Z\u0900-\u097F\s]*)$', line)
    if match:
        return (match.group(2).strip(), match.group(1).strip())

    # Format 5: "Bengan10" — text pehle number baad mein
    match = re.match(r'^([\w\u0900-\u097F\s]+?)\s*(\d+\.?\d*\s*(?:kg|gm|gram|nag|gaddi|bundle|pcs|pc|nos)?)$', line, re.IGNORECASE)
    if match:
        item = match.group(1).strip()
        qty = match.group(2).strip()
        if item and qty:
            return (item, qty)

    return None

def parse_order(message):
    orders = []
    lines = message.strip().split("\n")
    for line in lines:
        result = parse_line(line)
        if result:
            orders.append(result)
    return orders

def format_qty_reply(kg, pieces, bundles):
    parts = []
    if kg:
        parts.append(f"{kg} kg")
    if pieces:
        parts.append(f"{pieces} pcs")
    if bundles:
        parts.append(f"{bundles} bundle")
    return " | ".join(parts) if parts else "?"

@app.route("/webhook", methods=["POST"])
def webhook():
    incoming_msg = request.values.get("Body", "").strip()
    sender = request.values.get("From", "").replace("whatsapp:", "")
    media_url = request.values.get("MediaUrl0", "")
    account_sid = request.values.get("AccountSid", "")
    auth_token = os.environ.get("TWILIO_AUTH_TOKEN")

    print(f"📩 Message from {sender}:\n{incoming_msg}")

    resp = MessagingResponse()
    msg = resp.message()
    name = get_name(sender)

    if media_url:
        print(f"📷 Image mili!")
        gemini_text = extract_orders_from_image(media_url, account_sid, auth_token)
        print(f"Gemini response: {gemini_text}")
        if gemini_text:
            image_orders = parse_gemini_image_response(gemini_text)
            if image_orders:
                reply_lines = [f"📷 Image Orders Received! ({name})"]
                for item, kg, pieces, bundles in image_orders:
                    log_order(name, sender, item, kg, pieces, bundles)
                    qty_display = format_qty_reply(kg, pieces, bundles)
                    reply_lines.append(f"• {item} — {qty_display}")
                reply_lines.append("Thank you! 🙏")
                msg.body("\n".join(reply_lines))
                return str(resp)
        msg.body("❌ Image se orders detect nahi hue! Please clear image bhejo.")
        return str(resp)

    orders = parse_order(incoming_msg)
    if orders:
        reply_lines = [f"✅ Orders Received! ({name})"]
        for item, quantity in orders:
            english_item = translate_item(item)
            kg, pieces, bundles = parse_quantity(quantity)
            log_order(name, sender, english_item, kg, pieces, bundles)
            qty_display = format_qty_reply(kg, pieces, bundles)
            reply_lines.append(f"• {english_item.title()} — {qty_display}")
        reply_lines.append("Thank you! 🙏")
        msg.body("\n".join(reply_lines))
    else:
        msg.body("❌ Format samajh nahi aaya!\nExample:\nTamatar - 10kg\nDhaniya - 500gm\nLemon - 2 nag\nPudina - 1 gaddi")

    return str(resp)

if __name__ == "__main__":
    setup_excel()
    print("🤖 Bot server shuru ho gaya!")
    print("🌐 Webhook: http://localhost:5000/webhook")
    app.run(debug=False, port=5000)